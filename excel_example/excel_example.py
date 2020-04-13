import openpyxl as xl
from openpyxl.chart import Reference, BarChart

wb = xl.load_workbook("../files/excel/sample.xlsx")
sheet = wb["items"]

for row in range(2, sheet.max_row + 1):
    priceCell = sheet.cell(row, 2)
    descountCell = sheet.cell(row, 3)
    totalCell = sheet.cell(row, 4)

    totalCell.value = priceCell.value - ((descountCell.value * priceCell.value) / 100)

reference = Reference(sheet, min_row=2, max_row=5, min_col=1, max_col=4)
chart = BarChart()
chart.add_data(reference)

sheet.add_chart(chart, "A10")

wb.save("sample1.xlsx")

